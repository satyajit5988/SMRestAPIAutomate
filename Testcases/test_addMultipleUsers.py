import json
import jsonpath
import openpyxl
import requests


def test_Add_Multiple_Users():
    API_URL="https://thetestingworldapi.com/api/studentsDetails"
    file = open("C:/Users/User/PycharmProjects/SM - RestAPIAutomate/TestData/addMultipleUserDetails.json", "r")
    request_json = json.loads(file.read())

    wk = openpyxl.load_workbook("C:/Users/User/PycharmProjects/SM - RestAPIAutomate/TestData/userMultipleData.xlsx")
    sh = wk['Sheet1']
    row = sh.max_row

    for i in range(2,row+1):
        cell_first_name = sh.cell(row=i, column=1 )
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)

        request_json['first_name'] = cell_first_name.value
        request_json['middle_name'] = cell_mid_name.value
        request_json['last_name'] = cell_last_name.value
        request_json['date_of_birth'] = cell_dob.value

        response = requests.post(API_URL, request_json)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201